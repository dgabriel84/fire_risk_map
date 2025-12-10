import pandas as pd
import numpy as np
import geopandas as gpd
from shapely.geometry import Point, Polygon
import requests
from datetime import datetime, timedelta, timezone
from io import StringIO
from geopy.distance import great_circle
import streamlit as st

# --- FUNCIONES DE CARGA, SIMULACIÓN Y ANÁLISIS ---

@st.cache_data
def cargar_datos_instalaciones(file_path_or_buffer):
    """Carga, limpia y prepara los datos en un GeoDataFrame a partir de un archivo."""
    if file_path_or_buffer is None:
        return gpd.GeoDataFrame()

    try:
        # Intentar leer con pandas detectando separador automáticamente o probando comunes
        try:
            df = pd.read_csv(file_path_or_buffer)
            # Si solo tiene una columna, probablemente el separador está mal
            if len(df.columns) < 2:
                file_path_or_buffer.seek(0)
                df = pd.read_csv(file_path_or_buffer, sep=';')
        except:
            # Si falla, intentar explícitamente con punto y coma
            if hasattr(file_path_or_buffer, 'seek'):
                file_path_or_buffer.seek(0)
            df = pd.read_csv(file_path_or_buffer, sep=';')

        # Normalizar nombres de columnas (eliminar espacios extra)
        df.columns = df.columns.str.strip()

        df = df.rename(columns={
            'LATITUD': 'LAT', 'LONGITUD': 'LON', 'EQUIPO': 'TIPO_EQUIPO',
            'POBLACIÓN': 'POBLACION', 'Nº INVENTARIO': 'ID_INSTALACION',
            'DENOMINACIÓN': 'NAME', 'NAME': 'NOMBRE_CORTO_ORIGINAL',
            'MANTENEDOR': 'MANTENEDOR', 'CLIENTES': 'CLIENTES',
            'EMPLAZAMIENTO': 'EMPLAZAMIENTO',
            'GP': 'PROVINCIA'
        })

        required_cols = ['LAT', 'LON', 'CLIENTES', 'ID_INSTALACION', 'MANTENEDOR', 'EMPLAZAMIENTO', 'POBLACION', 'NAME',
                         'TIPO_EQUIPO', 'PROVINCIA']
        
        # Check for missing columns but don't stop, just return empty if critical ones are missing
        if not all(col in df.columns for col in required_cols):
             # In a real app we might want to raise an exception or return specific error info
             # For now we will try to proceed if LAT/LON exist
             pass

        if 'CCAA' not in df.columns: df['CCAA'] = 'N/A'

        # Manejar decimales con coma
        if df['LAT'].dtype == object:
            df['LAT'] = df['LAT'].astype(str).str.replace(',', '.').astype(float)
        if df['LON'].dtype == object:
            df['LON'] = df['LON'].astype(str).str.replace(',', '.').astype(float)

        df['LAT'] = pd.to_numeric(df['LAT'], errors='coerce')
        df['LON'] = pd.to_numeric(df['LON'], errors='coerce')
        df['CLIENTES'] = df['CLIENTES'].fillna(0).astype(int)
        df.dropna(subset=['LAT', 'LON'], inplace=True)

        if df.empty:
            return gpd.GeoDataFrame()

        # Crear URLs
        df['GMAPS_URL'] = "https://www.google.com/maps/search/?api=1&query=" + df['LAT'].astype(str) + "," + df['LON'].astype(str)
        df['STREET_URL'] = "https://www.google.com/maps/@?api=1&map_action=pano&viewpoint=" + df['LAT'].astype(str) + "," + df['LON'].astype(str)
        
        df['TOOLTIP_INFO'] = (
                "<b><font size='+1'>" + df['NAME'].astype(str) + "</font></b><br>"
                "<b>MANTENEDOR:</b> " + df['MANTENEDOR'].astype(str) + "<br>"
                "<b>CLIENTES:</b> " + df['CLIENTES'].astype(str) + "<br>"
                "<b>EMPLAZAMIENTO:</b> " + df['EMPLAZAMIENTO'].astype(str) + "<br><br>"
                "<b>🗺️ Google Maps:</b><br>"
                "<input type='text' value='" + df['GMAPS_URL'] + "' readonly style='width:100%;margin-bottom:8px;' onclick='this.select()'/><br>"
                "<b>🚶 Street View:</b><br>"
                "<input type='text' value='" + df['STREET_URL'] + "' readonly style='width:100%;' onclick='this.select()'/>"
        )

        gdf = gpd.GeoDataFrame(
            df,
            geometry=gpd.points_from_xy(df.LON, df.LAT),
            crs="EPSG:4326"
        )

        return gdf

    except Exception as e:
        print(f"Error al procesar el archivo CSV: {str(e)}")
        return gpd.GeoDataFrame()


@st.cache_data(ttl=3600) # Cache for 1 hour
def obtener_incendios_nasa_real(map_key, _gdf_instalaciones=None, area_bbox=None, dias=1, modo_prueba=False):
    """
    Conexión API de NASA FIRMS.
    """

    if not map_key and not modo_prueba:
        return pd.DataFrame()

    if modo_prueba:
        if _gdf_instalaciones is None or _gdf_instalaciones.empty:
            return pd.DataFrame()

        instalaciones_lat = _gdf_instalaciones['LAT'].mean()
        instalaciones_lon = _gdf_instalaciones['LON'].mean()

        n_incendios = 8
        np.random.seed(int(datetime.now().timestamp()))

        data = {
            'ID_INCENDIO': [f"TEST_FIRE_{i}" for i in range(n_incendios)],
            'LAT': instalaciones_lat + np.random.uniform(-0.5, 0.5, n_incendios),
            'LON': instalaciones_lon + np.random.uniform(-0.5, 0.5, n_incendios),
            'BRIGHTNESS': np.random.uniform(320, 380, n_incendios).round(2),
            'CONFIDENCE': np.random.randint(60, 95, n_incendios),
            'FECHA': datetime.now().strftime('%Y-%m-%d'),
            'HORA': '1200'
        }

        df_test = pd.DataFrame(data)
        
        df_test['FIRMS_URL'] = df_test.apply(
            lambda row: f"https://firms.modaps.eosdis.nasa.gov/map/#d:{row['FECHA']};@{row['LON']:.4f},{row['LAT']:.4f},14z",
            axis=1
        )

        return df_test

    # Modo real
    if area_bbox is None:
        # Usar siempre BBOX de España extendido para consistencia
        # [min_lon, min_lat, max_lon, max_lat]
        # Cubre Península, Baleares, Canarias y Ceuta/Melilla con margen
        # BBOX Península Ibérica (Aprox)
        # Excluye Canarias (Lon < -10)
        area_bbox = [-9.6, 35.9, 3.5, 44.0]

    area_str = f"{area_bbox[0]},{area_bbox[1]},{area_bbox[2]},{area_bbox[3]}"
    base_url = "https://firms.modaps.eosdis.nasa.gov/api/area/csv"
    
    # Fuentes a consultar (Multi-Satélite)
    sources = ["VIIRS_SNPP_NRT", "VIIRS_NOAA20_NRT"]
    dfs = []

    for source in sources:
        url = f"{base_url}/{map_key}/{source}/{area_str}/{dias}"
        try:
            response = requests.get(url, timeout=30)
            if response.status_code == 200:
                df_temp = pd.read_csv(StringIO(response.text))
                if not df_temp.empty:
                    dfs.append(df_temp)
            else:
                print(f"Error al conectar con NASA FIRMS ({source}). Código: {response.status_code}")
        except Exception as e:
            print(f"Error inesperado al cargar incendios ({source}): {str(e)}")

    if not dfs:
        return pd.DataFrame()

    df_incendios = pd.concat(dfs, ignore_index=True)

    if df_incendios.empty:
        return pd.DataFrame()

    df_incendios = df_incendios.rename(columns={
        'latitude': 'LAT',
        'longitude': 'LON',
        'bright_ti4': 'BRIGHTNESS',
        'confidence': 'CONFIDENCE',
        'acq_date': 'FECHA',
        'acq_time': 'HORA'
    })

    df_incendios['ID_INCENDIO'] = [f"FIRE_{i}" for i in range(len(df_incendios))]

    df_incendios['LAT'] = pd.to_numeric(df_incendios['LAT'], errors='coerce')
    df_incendios['LON'] = pd.to_numeric(df_incendios['LON'], errors='coerce')
    df_incendios['BRIGHTNESS'] = pd.to_numeric(df_incendios['BRIGHTNESS'], errors='coerce')

    # Mapeo de Confianza VIIRS Correcto
    if df_incendios['CONFIDENCE'].dtype == 'object':
        confidence_map = {
            'l': 15,   # Ruido probable
            'n': 85,   # Fuego muy probable
            'h': 100   # Fuego seguro / Saturación
        }
        df_incendios['CONFIDENCE'] = df_incendios['CONFIDENCE'].map(
            lambda x: confidence_map.get(str(x).lower(), 50)
        )

    df_incendios.dropna(subset=['LAT', 'LON'], inplace=True)
    
    df_incendios['FIRMS_URL'] = df_incendios.apply(
        lambda row: f"https://firms.modaps.eosdis.nasa.gov/map/#d:{row['FECHA']};@{row['LON']:.4f},{row['LAT']:.4f},14z",
        axis=1
    )

    # --- FILTRADO GEOGRÁFICO ---
    # BBOX Península Ibérica Estricto
    # [min_lon, min_lat, max_lon, max_lat]
    area_bbox = [-9.6, 35.9, 3.5, 44.0]
    
    # Filtrar por BBOX estricto
    df_incendios = df_incendios[
        (df_incendios['LAT'] >= area_bbox[1]) & (df_incendios['LAT'] <= area_bbox[3]) &
        (df_incendios['LON'] >= area_bbox[0]) & (df_incendios['LON'] <= area_bbox[2])
    ]

    # Excluir Baleares explícitamente (Lon > 0.9 y Lat < 41.0)
    # Delta del Ebro está aprox en 0.7E, 40.7N (Se mantiene)
    # Ibiza empieza en 1.2E (Se elimina)
    df_incendios = df_incendios[~((df_incendios['LON'] > 0.9) & (df_incendios['LAT'] < 41.0))]

    return df_incendios




@st.cache_data(show_spinner="⏳ Calculando riesgo por proximidad...", ttl=3600)
def calcular_riesgo_por_proximidad(_gdf_instalaciones, df_incendios, distancia_max_m):
    """Calcula el riesgo de proximidad utilizando Haversine vectorizado.
    
    Args:
        _gdf_instalaciones: GeoDataFrame con instalaciones (prefijo _ para no hashear)
        df_incendios: DataFrame con incendios
        distancia_max_m: Distancia máxima en metros para considerar riesgo
    
    Returns:
        GeoDataFrame con columnas de riesgo calculadas
    """
    gdf_result = _gdf_instalaciones.copy()
    gdf_result['COMPROMETIDA_DISTANCIA'] = False
    gdf_result['DISTANCIA_MIN_M'] = np.inf
    gdf_result['MAX_CONFIDENCE'] = 0

    if df_incendios.empty or gdf_result.empty:
        return gdf_result

    # Convertir a radianes
    lat1 = np.radians(gdf_result['LAT'].values)
    lon1 = np.radians(gdf_result['LON'].values)
    lat2 = np.radians(df_incendios['LAT'].values)
    lon2 = np.radians(df_incendios['LON'].values)
    confidences = df_incendios['CONFIDENCE'].fillna(0).values

    # Radio de la Tierra en metros
    R = 6371000

    # Inicializar arrays de resultados
    min_dists = np.full(len(gdf_result), np.inf)
    max_confs = np.zeros(len(gdf_result))
    comprometidas = np.zeros(len(gdf_result), dtype=bool)

    # Procesamiento por bloques para evitar uso excesivo de memoria si hay muchos datos
    # Aunque con numpy broadcasting puro suele ser eficiente para N*M < 10^8
    # N instalaciones ~ 10k, M incendios ~ 1k -> 10^7 operaciones, manejable.
    
    # Broadcasting: (N, 1) vs (1, M) -> (N, M)
    dlat = lat2[np.newaxis, :] - lat1[:, np.newaxis]
    dlon = lon2[np.newaxis, :] - lon1[:, np.newaxis]

    a = np.sin(dlat / 2)**2 + np.cos(lat1[:, np.newaxis]) * np.cos(lat2[np.newaxis, :]) * np.sin(dlon / 2)**2
    c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
    dist_matrix = R * c

    # Encontrar distancia mínima por instalación
    min_dists = np.min(dist_matrix, axis=1)
    
    # Máscara de incendios dentro del rango por instalación
    within_range_mask = dist_matrix <= distancia_max_m
    
    # Calcular Max Confidence solo para los que están en rango
    # Esto es un poco más complejo vectorizado sin loop, pero podemos hacerlo:
    # Multiplicamos la máscara por la confianza. Si no está en rango, es 0.
    # Luego tomamos el máximo por fila.
    
    # Expandir confianzas a matriz (1, M) -> (N, M)
    conf_matrix = np.tile(confidences, (len(gdf_result), 1))
    
    # Aplicar máscara: donde no está en rango, ponemos 0
    masked_conf = np.where(within_range_mask, conf_matrix, 0)
    
    max_confs = np.max(masked_conf, axis=1)
    comprometidas = min_dists <= distancia_max_m

    gdf_result['DISTANCIA_MIN_M'] = min_dists
    gdf_result['COMPROMETIDA_DISTANCIA'] = comprometidas
    gdf_result['MAX_CONFIDENCE'] = max_confs

    return gdf_result



@st.cache_data(show_spinner="⏳ Calculando riesgo por polígono...", ttl=3600)
def calcular_riesgo_por_poligono(_gdf_instalaciones, poligono_coords_tuple):
    """Calcula si una instalación está dentro del polígono dibujado.
    
    Args:
        _gdf_instalaciones: GeoDataFrame con instalaciones (prefijo _ para no hashear)
        poligono_coords_tuple: Tupla de coordenadas del polígono (hasheable)
    
    Returns:
        GeoDataFrame con columna COMPROMETIDA_POLIGONO
    """
    gdf_result = _gdf_instalaciones.copy()

    if gdf_result.geometry.is_empty.all() or gdf_result.geometry.crs is None:
        gdf_result['geometry'] = gpd.points_from_xy(gdf_result.LON, gdf_result.LAT)
        gdf_result = gpd.GeoDataFrame(gdf_result, geometry='geometry', crs="EPSG:4326")

    if 'COMPROMETIDA_POLIGONO' not in gdf_result.columns:
        gdf_result['COMPROMETIDA_POLIGONO'] = False

    if not poligono_coords_tuple:
        return gdf_result

    try:
        # Convertir tupla de vuelta a lista de coordenadas
        poligono_coords = list(poligono_coords_tuple)
        
        # poligono_coords espera ser una lista de tuplas (lat, lon).
        # Folium devuelve (lon, lat), pero el frontend ya lo ha convertido a (lat, lon).
        # Shapely espera (lon, lat), por lo que aquí invertimos (lat, lon) -> (lon, lat).
        # WAIT: If frontend sends (lat, lon), and we do [(lon, lat) for lat, lon in poligono_coords]
        # Then we are creating (lon, lat) tuples. This is CORRECT for Shapely.
        # Let's verify if the point.within check expects (lon, lat) or (lat, lon).
        # GeoPandas points are (lon, lat) if created from_xy(lon, lat).
        # So Polygon must be (lon, lat).
        # The logic seems correct.
        # Let's add a debug print to see what's happening.
        
        poly_points = [(lon, lat) for lat, lon in poligono_coords]
        poligono_shapely = Polygon(poly_points)
        
        # Ensure CRS match? GeoPandas handles it if geometry is set.
        # But here we are using raw shapely geometry against gdf geometry.
        # gdf geometry is (lon, lat).
    except Exception as e:
        print(f"Error creating polygon: {e}")
        return gdf_result
    except Exception:
        return gdf_result

    gdf_result['COMPROMETIDA_POLIGONO'] = gdf_result.geometry.apply(
        lambda point: point.within(poligono_shapely)
    )

    return gdf_result



def generar_informe_compromiso(gdf_riesgo):
    """Genera un DataFrame para el informe."""
    if gdf_riesgo.empty:
        return None

    df_comprometidas = gdf_riesgo[gdf_riesgo['COMPROMETIDA_FINAL']].copy()

    if df_comprometidas.empty:
        return None

    df_informe = df_comprometidas.sort_values(by='DISTANCIA_MIN_M', ascending=True).copy()

    df_informe['DISTANCIA_MIN_KM'] = (df_informe['DISTANCIA_MIN_M'] / 1000).round(2)
    df_informe['DISTANCIA_MIN_KM'] = df_informe['DISTANCIA_MIN_KM'].apply(
        lambda x: "N/A (Riesgo Manual)" if x == np.inf else x
    )

    df_informe = df_informe.rename(columns={
        'ID_INSTALACION': 'Nº INVENTARIO',
    })

    columnas_finales = [
        'Nº INVENTARIO',
        'POBLACION',
        'PROVINCIA',
        'CCAA',
        'CLIENTES',
        'EMPLAZAMIENTO',
        'MANTENEDOR',
        'DISTANCIA_MIN_KM'
    ]
    
    # Filter only existing columns
    cols_to_use = [c for c in columnas_finales if c in df_informe.columns]
    df_informe = df_informe[cols_to_use]

    return df_informe


def filtrar_incendios_por_tiempo(df, filtro):
    """Filtra el DataFrame de incendios según la antigüedad seleccionada."""
    if df.empty or filtro == "Todas":
        return df

    try:
        horas_max = int(filtro.split()[0])
    except ValueError:
        return df

    now_utc = datetime.now(timezone.utc)

    def es_reciente(row):
        try:
            fecha_str = row.get('FECHA', now_utc.strftime('%Y-%m-%d'))
            hora_str = str(row.get('HORA', '0000')).zfill(4)
            dt_str = f"{fecha_str} {hora_str}"
            
            dt_fire = datetime.strptime(dt_str, "%Y-%m-%d %H%M").replace(tzinfo=timezone.utc)
            
            diff = now_utc - dt_fire
            return timedelta(hours=0) <= diff <= timedelta(hours=horas_max)
        except Exception:
            return False

    return df[df.apply(es_reciente, axis=1)].copy()


def generar_informe_excel(gdf_riesgo, df_incendios, filtros_info):
    """Genera un archivo Excel con múltiples hojas e hipervínculos."""
    if gdf_riesgo.empty and df_incendios.empty:
        return None

    output = StringIO() # Not used for Excel, we return the object or save it
    # We will return a BytesIO object or let the GUI handle the saving if we pass a path.
    # Better: return a pandas ExcelWriter object or just the dataframes to be written?
    # GUI expects to save to a file. Let's return None and let GUI handle file dialog, 
    # then we receive the path here? No, better to separate logic.
    # Let's create the Excel in memory or return a function that writes to a path.
    pass

def guardar_excel_completo(filepath, gdf_riesgo, df_incendios, filtros_info):
    """Guarda el informe completo en Excel en la ruta especificada."""
    
    # 1. Preparar Datos de Instalaciones
    df_inst = pd.DataFrame()
    if not gdf_riesgo.empty and 'COMPROMETIDA_FINAL' in gdf_riesgo.columns:
        df_inst = gdf_riesgo[gdf_riesgo['COMPROMETIDA_FINAL']].copy()
        if not df_inst.empty:
            df_inst = df_inst.sort_values(by='DISTANCIA_MIN_M')
            df_inst['DISTANCIA_KM'] = (df_inst['DISTANCIA_MIN_M'] / 1000).round(2)
            df_inst['DISTANCIA_KM'] = df_inst['DISTANCIA_KM'].replace(np.inf, "Polígono")
            
            # Asegurar que MAX_CONFIDENCE existe
            if 'MAX_CONFIDENCE' not in df_inst.columns:
                df_inst['MAX_CONFIDENCE'] = 0
            
            cols_export = [
                'ID_INSTALACION', 'NAME', 'POBLACION', 'PROVINCIA', 'CCAA', 
                'CLIENTES', 'MANTENEDOR', 'EMPLAZAMIENTO', 'TIPO_EQUIPO', 'DISTANCIA_KM',
                'MAX_CONFIDENCE', 'GMAPS_URL', 'STREET_URL'
            ]
            # Filter existing columns
            cols_export = [c for c in cols_export if c in df_inst.columns]
            df_inst = df_inst[cols_export]

    # 2. Preparar Datos de Incendios
    df_fire = pd.DataFrame()
    if not df_incendios.empty:
        df_fire = df_incendios.copy()
        cols_fire = ['ID_INCENDIO', 'LAT', 'LON', 'FECHA', 'HORA', 'BRIGHTNESS', 'CONFIDENCE', 'FIRMS_URL']
        cols_fire = [c for c in cols_fire if c in df_fire.columns]
        df_fire = df_fire[cols_fire]

    # 3. Preparar Resumen
    total_clientes = 0
    if not df_inst.empty and 'CLIENTES' in df_inst.columns:
        total_clientes = df_inst['CLIENTES'].sum()

    resumen_data = {
        'Fecha Informe': [datetime.now().strftime('%d/%m/%Y %H:%M')],
        'Total Instalaciones Afectadas': [len(df_inst)],
        'Total Clientes Afectados': [total_clientes],
        'Total Incendios (Península)': [len(df_fire)],
        '--- FILTROS ---': [''],
    }
    for k, v in filtros_info.items():
        resumen_data[k] = [str(v)]
    
    df_resumen = pd.DataFrame.from_dict(resumen_data, orient='index', columns=['Valor'])
    df_resumen.reset_index(inplace=True)
    df_resumen.columns = ['Parámetro', 'Valor']

    # 4. Escribir Excel
    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Hoja Resumen
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja Instalaciones
            if not df_inst.empty:
                df_inst.to_excel(writer, sheet_name='Instalaciones Afectadas', index=False)
                # Añadir hipervínculos y formato condicional
                wb = writer.book
                ws = wb['Instalaciones Afectadas']
                
                # Definir estilos
                from openpyxl.styles import PatternFill, Font
                red_font = Font(color="FF0000", bold=True)
                
                fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Nominal (85)
                fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # High (100)
                
                # Tipos a resaltar (Emplazamiento)
                tipos_criticos = [
                    "AÉREAS (A)", "Armar.rec.vall (AV)", "Armario (AR)", 
                    "Arqueta Pared (PA)", "AZOTEA (AZ)", "CASETA (CA)"
                ]
                
                # Encontrar índices de columnas
                idx_gmaps = -1
                idx_street = -1
                idx_tipo = -1
                idx_empl = -1
                idx_conf = -1
                
                cols = {c: i+1 for i, c in enumerate(df_inst.columns)}
                idx_gmaps = cols.get('GMAPS_URL', -1)
                idx_street = cols.get('STREET_URL', -1)
                idx_tipo = cols.get('TIPO_EQUIPO', -1)
                idx_empl = cols.get('EMPLAZAMIENTO', -1)
                idx_conf = cols.get('MAX_CONFIDENCE', -1)

                for row_idx in range(2, len(df_inst) + 2): # Skip header
                    # 1. Colorear Fila y Formatear Texto según Confianza
                    if idx_conf > 0:
                        conf_val = ws.cell(row=row_idx, column=idx_conf).value
                        try:
                            conf_num = float(conf_val)
                            
                            # Determinar Texto
                            conf_text = f"{int(conf_num)}%"
                            if conf_num == 15:
                                conf_text = "Baja (15%)"
                            elif conf_num == 85:
                                conf_text = "Nominal (85%)"
                            elif conf_num >= 100:
                                conf_text = "Alta (100%)"
                            
                            # Escribir Texto
                            ws.cell(row=row_idx, column=idx_conf).value = conf_text
                            
                            # Aplicar Color (basado en valor numérico original)
                            if conf_num >= 100:
                                for col_idx in range(1, len(df_inst.columns) + 1):
                                    ws.cell(row=row_idx, column=col_idx).fill = fill_red
                            elif conf_num >= 85:
                                for col_idx in range(1, len(df_inst.columns) + 1):
                                    ws.cell(row=row_idx, column=col_idx).fill = fill_orange
                                    
                        except (ValueError, TypeError):
                            pass

                    # 2. Resaltar Texto Emplazamiento/Tipo si es crítico
                    # Check both TIPO_EQUIPO and EMPLAZAMIENTO
                    is_critical = False
                    if idx_tipo > 0:
                        val = str(ws.cell(row=row_idx, column=idx_tipo).value or '').strip()
                        if val in tipos_criticos: is_critical = True
                    if idx_empl > 0:
                        val = str(ws.cell(row=row_idx, column=idx_empl).value or '').strip()
                        if val in tipos_criticos: is_critical = True
                    
                    if is_critical and idx_empl > 0:
                         ws.cell(row=row_idx, column=idx_empl).font = red_font

                    # Hipervínculos
                    if idx_gmaps > 0:
                        cell = ws.cell(row=row_idx, column=idx_gmaps)
                        if cell.value:
                            cell.hyperlink = cell.value
                            cell.value = "Ver Mapa"
                            cell.style = "Hyperlink"
                    if idx_street > 0:
                        cell = ws.cell(row=row_idx, column=idx_street)
                        if cell.value:
                            cell.hyperlink = cell.value
                            cell.value = "Ver Street View"
                            cell.style = "Hyperlink"
            
            # Hoja Incendios
            if not df_fire.empty:
                df_fire.to_excel(writer, sheet_name='Incendios NASA', index=False)
                wb = writer.book
                ws = wb['Incendios NASA']
                
                idx_firms = -1
                if 'FIRMS_URL' in df_fire.columns:
                    idx_firms = df_fire.columns.get_loc('FIRMS_URL') + 1
                
                for row_idx in range(2, len(df_fire) + 2):
                    if idx_firms > 0:
                        cell = ws.cell(row=row_idx, column=idx_firms)
                        if cell.value:
                            cell.hyperlink = cell.value
                            cell.value = "Ver FIRMS"
                            cell.style = "Hyperlink"
                            
        return True
    except Exception as e:
        print(f"Error guardando Excel: {e}")
        return False

def generar_html_report(gdf_riesgo, df_incendios, filtros_info):
    """Genera el contenido HTML para el informe PDF."""
    now = datetime.now()
    
    # Preparar datos
    df_inst = pd.DataFrame()
    if not gdf_riesgo.empty and 'COMPROMETIDA_FINAL' in gdf_riesgo.columns:
        df_inst = gdf_riesgo[gdf_riesgo['COMPROMETIDA_FINAL']].copy()
        if not df_inst.empty:
            df_inst = df_inst.sort_values(by='DISTANCIA_MIN_M')
    
    df_fire = df_incendios.copy() if not df_incendios.empty else pd.DataFrame()

    # Calcular total clientes
    total_clientes = 0
    if not df_inst.empty and 'CLIENTES' in df_inst.columns:
        total_clientes = df_inst['CLIENTES'].sum()

    # Construir HTML
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Helvetica, Arial, sans-serif; font-size: 10pt; }}
            h1 {{ color: #2c3e50; text-align: center; }}
            h2 {{ color: #16a085; border-bottom: 2px solid #16a085; padding-bottom: 5px; margin-top: 20px; }}
            .info-box {{ background-color: #f8f9fa; border: 1px solid #ddd; padding: 10px; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 15px; font-size: 9pt; }}
            th {{ background-color: #2c3e50; color: white; padding: 8px; text-align: left; }}
            td {{ border: 1px solid #ddd; padding: 6px; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .critical-text {{ color: red; font-weight: bold; }}
            .row-warning {{ background-color: #ffeeba; }} 
            .row-danger {{ background-color: #f5c6cb; }}
        </style>
    </head>
    <body>
        <h1>Informe de instalaciones en riesgo</h1>
        <div class="info-box">
            <p><b>Fecha de Emisión:</b> {now.strftime('%d/%m/%Y %H:%M')}</p>
            <p><b>Total Clientes Afectados:</b> {total_clientes}</p>
            <p><b>Filtros Aplicados:</b></p>
            <ul>
    """
    for k, v in filtros_info.items():
        if v: html += f"<li><b>{k}:</b> {v}</li>"
    
    html += """
            </ul>
        </div>
    """

    # Tabla Instalaciones
    html += f"<h2>Instalaciones Afectadas ({len(df_inst)})</h2>"
    if not df_inst.empty:
        html += """
        <table>
            <tr>
                <th>ID</th>
                <th>Nombre</th>
                <th>Población</th>
                <th>Emplazamiento</th>
                <th>Clientes</th>
                <th>Distancia</th>
                <th>Confianza Max</th>
                <th>Enlaces</th>
            </tr>
        """
        
        tipos_criticos = [
            "AÉREAS (A)", "Armar.rec.vall (AV)", "Armario (AR)", 
            "Arqueta Pared (PA)", "AZOTEA (AZ)", "CASETA (CA)"
        ]

        for _, row in df_inst.iterrows():
            dist_km = row.get('DISTANCIA_MIN_M', 0) / 1000
            dist_str = f"{dist_km:.2f} km" if dist_km != float('inf') else "Polígono"
            
            tipo = str(row.get('TIPO_EQUIPO', '')).strip()
            empl = str(row.get('EMPLAZAMIENTO', '')).strip()
            
            is_critical_type = (tipo in tipos_criticos) or (empl in tipos_criticos)
            empl_style = 'class="critical-text"' if is_critical_type else ''
            
            max_conf = row.get('MAX_CONFIDENCE', 0)
            conf_text = f"{int(max_conf)}%"
            if max_conf == 15: conf_text = "Baja (15%)"
            elif max_conf == 85: conf_text = "Nominal (85%)"
            elif max_conf >= 100: conf_text = "Alta (100%)"

            row_class = ""
            if max_conf >= 100: row_class = "class='row-danger'"
            elif max_conf >= 85: row_class = "class='row-warning'"
            
            gmaps = row.get('GMAPS_URL', '#')
            street = row.get('STREET_URL', '#')
            
            links_html = f"<a href='{gmaps}'>Mapa</a> | <a href='{street}'>Street</a>"
            
            html += f"""
            <tr {row_class}>
                <td>{row.get('ID_INSTALACION', '')}</td>
                <td>{row.get('NAME', '')}</td>
                <td>{row.get('POBLACION', '')}</td>
                <td {empl_style}>{row.get('EMPLAZAMIENTO', '')}</td>
                <td>{row.get('CLIENTES', 0)}</td>
                <td>{dist_str}</td>
                <td>{conf_text}</td>
                <td>{links_html}</td>
            </tr>
            """
        html += "</table>"
    else:
        html += "<p>No hay instalaciones afectadas.</p>"

    # Tabla Incendios
    html += f"<h2>Incendios Detectados ({len(df_fire)})</h2>"
    if not df_fire.empty:
        html += """
        <table>
            <tr>
                <th>ID</th>
                <th>Fecha/Hora</th>
                <th>Confianza</th>
                <th>Ubicación</th>
                <th>Enlace</th>
            </tr>
        """
        for _, row in df_fire.iterrows():
            hora = str(row.get('HORA', '')).zfill(4)
            hora_fmt = f"{hora[:2]}:{hora[2:]}"
            
            firms = row.get('FIRMS_URL', '#')
            
            html += f"""
            <tr>
                <td>{row.get('ID_INCENDIO', '')}</td>
                <td>{row.get('FECHA', '')} {hora_fmt}</td>
                <td>{row.get('CONFIDENCE', '')}%</td>
                <td>{row.get('LAT', '')}, {row.get('LON', '')}</td>
                <td><a href='{firms}'>Ver FIRMS</a></td>
            </tr>
            """
        html += "</table>"
    
    html += "</body></html>"
    return html
